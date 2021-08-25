import openpyxl, datetime
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.workbook import workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import os


class createExcel:
    def __init__(self):
        print("Do you want to remove previous workbook and create a new one, or load a previous version and update it?\n1. Remove (if applicable) and create a new one.\n2. Load and update a previous version.")
        choice = input()
        if choice == "1":
            try:
                os.remove("Deployment_NAC_Policy_Info.xlsx")
            except:
                print("File not found, no removal necessary.")
            finally:
                self.workbook = Workbook()
        elif choice == "2":
            try: 
                self.workbook = load_workbook(filename="Deployment_NAC_Policy_Info.xlsx")
                self.sheetnames = self.workbook.sheetnames
                print(self.workbook.sheetnames)
            except:
                print("You chose to load a previous version, but it does not exist. Creating a new workbook.")
                self.workbook = Workbook()
        else:
            print("You did not make a valid choce, please restart and choose a valid option.")

    def saveWorkBook(self):
        if "Sheet" in self.workbook.sheetnames:
            blank_sheet = self.workbook.get_sheet_by_name('Sheet')
            self.workbook.remove_sheet(blank_sheet)
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~ SAVING WORKBOOK ~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        self.workbook.save("Deployment_NAC_Policy_Info.xlsx")

    
def authczConditionWorkbook(self,policy_set_info):
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~ WRITING AUTHC/Z DATA TO EXCEL ~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        if "Authc and Authz Rules" in self.workbook.sheetnames:
            ws = self.workbook.remove_sheet("Authc and Authz Rules")
        ws = self.workbook.create_sheet("Authc and Authz Rules")
        ws.insert_cols(0,100)
        cell_row = 1
        for policy_set in policy_set_info:
            ws.insert_rows(ws.max_row,2)
            ##Authentication policy
            ws["A{}".format(cell_row)] = "Policy Set: {}".format(policy_set)
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
            ws["A{}".format(cell_row)].font = Font(size="24",bold=True,color="FFFFFF")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws.merge_cells("A{}:B{}".format(cell_row,cell_row))
            cell_row += 1
            ws["A{}".format(cell_row)] = "Authentication Rules"
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
            ws["A{}".format(cell_row)].font = Font(size="20",bold=True,color="FFFFFF")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws.merge_cells("A{}:B{}".format(cell_row,cell_row))
            cell_row += 1
            for authc_rule in policy_set_info[policy_set][0]:
                ws.insert_rows(ws.max_row,1)
                ws["A{}".format(cell_row)] = "Authentication Rule {}".format(policy_set_info[policy_set][0][authc_rule]["rule"]["name"])
                ws["B{}".format(cell_row)] = "references the {} identity source sequence".format(policy_set_info[policy_set][0][authc_rule]["identitySourceName"])
                ws["C{}".format(cell_row)] = "If authc fails: {}".format(policy_set_info[policy_set][0][authc_rule]["ifAuthFail"])
                ws["D{}".format(cell_row)] = "If process fails: {}".format(policy_set_info[policy_set][0][authc_rule]["ifProcessFail"])
                ws["E{}".format(cell_row)] = "If user is not found: {}".format(policy_set_info[policy_set][0][authc_rule]["ifUserNotFound"])
                cell_row += 1
                ascii_start = 70
                cond_string = ""
                if policy_set_info[policy_set][0][authc_rule]["rule"]["condition"] is not "null":
                    if "children" in policy_set_info[policy_set][0][authc_rule]["rule"]["condition"]:
                        for child in policy_set_info[policy_set][0][authc_rule]["rule"]["condition"]["children"]:
                            if "children" in child:
                                for childs_child in child:
                                    cond_string += "If "
                                    #IF ITS A CUSTOM
                                    if "attributeName" in childs_child:
                                        if childs_child['isNegate'] is "false":
                                            cond_string += "Attribute Name: {}, Value: {} {} TRUE ".format(childs_child['attributeName'],childs_child['attributeValue'],childs_child['operator'])
                                        else:
                                            cond_string += "Attribute Name: {}, Value: {} {} NOT TRUE ".format(childs_child['attributeName'],childs_child['attributeValue'],childs_child['operator'])
                                    ##IF ITS DICTIONARY CONDITION OR SAVED CONDITION
                                    ## POSSIBLE TO ENHANCE THIS BY PULLING ALL DICTIONARY CONDITIONS
                                    else:
                                        if childs_child['isNegate'] is "false":
                                            cond_string += "Dictionary condition named {} is TRUE ".format(childs_child['attributeName'],childs_child['attributeValue'],childs_child['operator'])
                                        else:
                                            cond_string += "Dictionary condition named {} is NOT TRUE ".format(childs_child['attributeName'],childs_child['attributeValue'],childs_child['operator'])
                                    


                                    

                           if type(child) is not dict:
                               break

                    
                    for condition in policy_set_info[policy_set][0][authc_rule]["rule"]["condition"]

                ws["F{}".format(cell_row)] = "If user is not found: {}".format(policy_set_info[policy_set][0][authc_rule]["ifUserNotFound"])
                ws["G{}".format(cell_row)] = "If user is not found: {}".format(policy_set_info[policy_set][0][authc_rule]["ifUserNotFound"])
                ws["A{}".format(cell_row)] = "Policy Set: {}".format(policy_set)
                ws["A{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
                ws["A{}".format(cell_row)].font = Font(size="24",bold=True,color="FFFFFF")
                ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
                



    def createDACLWorkBook(self,dacl_info):
        print("~~~~~~~~~~~~~~~~~~~~~~~~~~~~ WRITING DACL DATA TO EXCEL ~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
        if "DACLs In Use" in self.workbook.sheetnames:
            ws = self.workbook.remove_sheet("DACLs In Use")
        ws = self.workbook.create_sheet("DACLs In Use")
        ws.insert_cols(0,1)
        dacl_ace_holder = []
        cell_row = 1
        for authz_info in dacl_info:
            dacl_ace_holder = authz_info["dacl_info"]["dacl"].split("\n")
            num_of_rows = 6+len(dacl_ace_holder)
            ws.insert_rows(ws.max_row,num_of_rows)
            ws["A{}".format(cell_row)] = "Policy Set --> Authz Rule:"
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True,color="FFFFFF")
            ws["B{}".format(cell_row)] = "{} ---> {}".format(authz_info["authz_info"]["usedInPolicySet"],authz_info["authz_info"]["usedInAuthzRule"])
            ws["B{}".format(cell_row)].font = Font(size="14",bold=True,color="FFFFFF")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "Authz Profile Used"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["B{}".format(cell_row)] = authz_info["authz_info"]["profileName"]
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="74BF4B",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="74BF4B",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "Authz Profile Access Type"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["B{}".format(cell_row)] = authz_info["authz_profile"]["AuthorizationProfile"]["accessType"]
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="FBAB2C",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="FBAB2C",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "dACL Name"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["B{}".format(cell_row)] = authz_info["authz_profile"]["AuthorizationProfile"]["daclName"]
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "dACL Type"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["B{}".format(cell_row)] = authz_info["dacl_info"]['daclType']
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ace_index = 1
            for ace in dacl_ace_holder:
                ws["A{}".format(cell_row)] = "ACE #{}".format(ace_index)
                ace_index += 1
                ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
                ws["B{}".format(cell_row)] = ace
                ws["A{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
                ws["B{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
                ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
                ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
                cell_row += 1
        ws.column_dimensions["A"].width = 33
        ws.column_dimensions["B"].width = 50

    def createIDWorkBook(self, id_seqs):
        if "ID Source Sequences In Use" in self.workbook.sheetnames:
            ws = self.workbook.remove_sheet("ID Source Sequences In Use")
        ws = self.workbook.create_sheet("ID Source Sequences In Use")
        ws.insert_cols(0,1)
        cell_row = 1
        for id_info in id_seqs:
            num_of_rows = 5+len(id_info["id_seq"]["IdStoreSequence"]["idSeqItem"])
            ws.insert_rows(ws.max_row,num_of_rows)
            ws["A{}".format(cell_row)] = "Policy Set --> Authc Rule:"
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True,color="FFFFFF")
            ws["B{}".format(cell_row)] = "{} ---> {}".format(id_info["policy_set_info"]["policy_set_name"], id_info["policy_set_info"]["authc_rule_name"])
            ws["B{}".format(cell_row)].font = Font(size="14",bold=True,color="FFFFFF")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="1E4471",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "Identity Source Sequence Used"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["B{}".format(cell_row)] = id_info["id_seq"]["IdStoreSequence"]["name"]
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="74BF4B",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="74BF4B",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "CAP Used (If in use)"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["B{}".format(cell_row)] = id_info["id_seq"]["IdStoreSequence"]["certificateAuthenticationProfile"]
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="FBAB2C",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="FBAB2C",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            cell_row += 1
            ws["A{}".format(cell_row)] = "ID Stores"
            ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
            ws["A{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
            ws["B{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
            ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
            ws.merge_cells("A{}:B{}".format(cell_row,cell_row))
            cell_row += 1
            for id_store in id_info["id_seq"]["IdStoreSequence"]["idSeqItem"]:
                ws["A{}".format(cell_row)] = "ID Store #{}".format(id_store["order"])
                ws["A{}".format(cell_row)].font = Font(size="14",bold=True)
                ws["B{}".format(cell_row)] = id_store['idstore']
                ws["A{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
                ws["B{}".format(cell_row)].fill = PatternFill(fgColor="00BCEB",fill_type="solid")
                ws["A{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
                ws["B{}".format(cell_row)].border = Border(left=Side(border_style="thin",color='000000'),right=Side(border_style="thin",color='000000'),top=Side(border_style="thin",color='000000'),bottom=Side(border_style="thin",color='000000'))
                cell_row += 1
        ws.column_dimensions["A"].width = 33
        ws.column_dimensions["B"].width = 30


